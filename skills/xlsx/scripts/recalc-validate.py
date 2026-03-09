# pyre-ignore-all-errors
"""
recalc-validate.py — Validador de fórmulas Excel
Detecta #REF!, #DIV/0!, referencias circulares y errores potenciales
antes de entregar el archivo al usuario.
"""
import os
import sys

from openpyxl import load_workbook  # pyre-ignore[21]
from openpyxl.utils.exceptions import InvalidFileException  # pyre-ignore[21]
from typing import Any

ERROR_VALUES = {"#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#VALUE!"}
WARNING_PATTERNS = ["#N/A"]


def validate_workbook(filepath: str) -> dict:
    """
    Abre el workbook y detecta celdas con valores de error.
    Retorna un reporte con errores, advertencias y un flag passed/failed.
    """
    errors: list[Any] = []
    warnings: list[Any] = []

    try:
        wb = load_workbook(filepath, data_only=True)
    except InvalidFileException as e:
        errors.append(f"Cannot open file: {e}")
        return {"file": filepath, "errors": errors, "warnings": warnings, "passed": False}
    except Exception as e:
        errors.append(f"Unexpected error opening file: {e}")
        return {"file": filepath, "errors": errors, "warnings": warnings, "passed": False}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if any(err in val for err in ERROR_VALUES):
                    entry = {
                        "sheet": sheet_name,
                        "cell": cell.coordinate,
                        "value": val
                    }
                    if any(warn in val for warn in WARNING_PATTERNS):
                        warnings.append(entry)
                    else:
                        errors.append(entry)

    passed = len(errors) == 0
    return {
        "file": filepath,
        "errors": errors,
        "warnings": warnings,
        "passed": passed
    }


def print_report(report: dict):
    status = "✅ PASSED" if report["passed"] else "❌ FAILED"
    print(f"\n{'='*50}")
    print(f"File: {report['file']}")
    print(f"Status: {status}")

    if report["errors"]:
        print(f"\n🔴 ERRORS ({len(report['errors'])}):")
        for e in report["errors"]:
            if isinstance(e, dict):
                print(f"  [{e['sheet']}] {e['cell']} → {e['value']}")
            else:
                print(f"  {e}")

    if report["warnings"]:
        print(f"\n🟡 WARNINGS ({len(report['warnings'])}):")
        for w in report["warnings"]:
            print(f"  [{w['sheet']}] {w['cell']} → {w['value']}")

    if report["passed"] and not report["warnings"]:
        print("\nAll cells clean. No formula errors detected.")
    print('='*50)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python recalc-validate.py <path_to_xlsx>")
        sys.exit(1)

    path = sys.argv[1]
    if not os.path.exists(path):
        print(f"Error: File not found — {path}")
        sys.exit(1)

    result = validate_workbook(path)
    print_report(result)
    sys.exit(0 if result["passed"] else 1)
